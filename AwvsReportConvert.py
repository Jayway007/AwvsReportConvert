# -*- coding:utf-8 -*-
from bs4 import BeautifulSoup
import os
import re
import openpyxl as ws

def endWith(file,*endstring):
    array = map(file.endswith,endstring)
    if True in array:
        return True
    else:
        return False

def openFile():
    file = os.listdir('.')
    for k in range(len(file)):
        if endWith(file[k],'.html'):
            soup = BeautifulSoup(open(file[k], mode='r', encoding='utf-8'), 'lxml')
            get_detail(soup)
            
def has_border_but_no_class(tag):
    return tag.has_attr('border') and not tag.has_attr('class')

def get_detail(soup):
    vulnerabilities = {}
    tables = soup.find_all(has_border_but_no_class)
    url = soup.select(".ax-scan-summary > tbody > tr:nth-of-type(3) > td:nth-of-type(2)")[0].string
    len(list(tables))
    for table in tables:
        scan_url = url
        vl_path = table.select('tr > td > b')[0].string.strip() 
        vl_name = table.select('tr:nth-of-type(2) > td > b')[1].string.strip()
        vl_severity = table.select('tr:nth-of-type(3) > td:nth-of-type(2)')[0].string.strip()
        vl_description = table.select('tr:nth-of-type(4) > td:nth-of-type(2)')[0].get_text().strip()
        vl_detail = table.select('tr:nth-of-type(7) > td:nth-of-type(2)')[0].get_text().strip()
        vl_post = table.select('tr:nth-of-type(8) > td')[0].string
        vl_recommendations = table.select('tr:nth-of-type(5) > td:nth-of-type(2)')[0].get_text().strip()
        vulnerabilities['url'] = scan_url
        vulnerabilities['path'] = vl_path
        vulnerabilities['name'] = vl_name
        vulnerabilities['severity'] = vl_severity
        vulnerabilities['vl_description'] = vl_description
        vulnerabilities['detail'] = vl_detail
        vulnerabilities['post'] = vl_post
        vulnerabilities['recommendtions'] = vl_recommendations
        write_xlsx(vulnerabilities)

def write_xlsx(vulnerabilities):
    wb = ws.load_workbook("AwvsReport.xlsx")
    sheet1 = wb['Sheet']
    num = sheet1.max_row
    sheet1.cell(row = num+1, column=1, value=vulnerabilities['url'])
    sheet1.cell(row = num+1,column = 2,value = vulnerabilities['name'])
    sheet1.cell(row = num+1,column = 3,value = vulnerabilities['path'])
    sheet1.cell(row = num+1,column = 4,value = vulnerabilities['severity'])
    sheet1.cell(row = num+1,column = 5,value = vulnerabilities['vl_description'])
    sheet1.cell(row=num + 1, column=6, value=vulnerabilities['detail'])
    sheet1.cell(row=num + 1, column=7, value=vulnerabilities['post'])
    sheet1.cell(row=num + 1, column=8, value=vulnerabilities['recommendtions'])
    wb.save("AwvsReport.xlsx")

def creat_xlsx():
    s = 0
    wb = ws.Workbook()
    ws1 = wb.active
    word=['风险目标','风险名称','风险地址','风险等级','风险描述','风险详细','风险请求','整改意见'] #风险参数
    for i in word:
        s = s + 1
        ws1.cell(row =1,column = s,value = i)
    wb.save("AwvsReport.xlsx")

def main():
    print("Please enter any key to continue")
    os.system("pause")
    creat_xlsx()
    openFile()
    print("Completed！")
    os.system("pause")

if __name__ == '__main__':
    main()